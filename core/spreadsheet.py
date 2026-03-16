"""
core/spreadsheet.py
Builds a color-coded .xlsx with three tabs: Job Listings, Summary, Search Config.
"""

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# Color palette
COLOR_HEADER_BG = "1F3864"
COLOR_HEADER_FG = "FFFFFF"
COLOR_REMOTE     = "E2EFDA"
COLOR_HYBRID     = "FFF2CC"
COLOR_ALT_ROW    = "F2F2F2"
COLOR_ACCENT     = "2E75B6"
COLOR_SECTION    = "D9E1F2"


def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_style(cell, text=""):
    cell.value = text
    cell.font = Font(bold=True, color=COLOR_HEADER_FG, size=11)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()


def _cell_style(cell, value="", fill_color=None, bold=False, wrap=False, align="left"):
    cell.value = value
    if fill_color:
        cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin_border()


class SpreadsheetBuilder:
    def __init__(self, cfg):
        self.output_dir = Path(cfg.get("output_dir", "output"))
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def build(self, jobs: list[dict], profile: dict) -> str:
        wb = openpyxl.Workbook()
        ws_jobs = wb.active
        ws_jobs.title = "Job Listings"
        ws_summary = wb.create_sheet("Summary")
        ws_config = wb.create_sheet("Search Config")

        self._build_listings(ws_jobs, jobs, profile)
        self._build_summary(ws_summary, jobs, profile)
        self._build_config(ws_config, profile)

        date_str = datetime.now().strftime("%Y-%m-%d")
        safe_name = profile["name"].replace(" ", "_")
        filename = f"{safe_name}_Jobs_{date_str}.xlsx"
        filepath = self.output_dir / filename
        wb.save(str(filepath))

        latest_path = self.output_dir / f"{safe_name}_Jobs_LATEST.xlsx"
        import shutil
        shutil.copy(str(filepath), str(latest_path))

        return str(filepath)

    def _build_listings(self, ws, jobs, profile):
        headers = ["#", "Job Title", "Company", "Location", "Salary",
                   "Work Type", "Source", "Date Found", "Query", "Apply Link"]
        col_widths = [5, 35, 25, 25, 20, 12, 14, 12, 30, 50]

        ws.row_dimensions[1].height = 30
        for col, (hdr, width) in enumerate(zip(headers, col_widths), 1):
            _header_style(ws.cell(1, col), hdr)
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = "A2"

        for row_idx, job in enumerate(jobs, 2):
            loc = job.get("location", "").lower()
            if "remote" in loc:
                fill = COLOR_REMOTE
            elif "hybrid" in loc:
                fill = COLOR_HYBRID
            else:
                fill = COLOR_ALT_ROW if row_idx % 2 == 0 else "FFFFFF"

            work_type = ("Remote" if "remote" in loc
                         else "Hybrid" if "hybrid" in loc else "On-site")

            values = [
                row_idx - 1,
                job.get("title", ""),
                job.get("company", ""),
                job.get("location", ""),
                job.get("salary", ""),
                work_type,
                job.get("source", "").capitalize(),
                job.get("date_found", ""),
                job.get("query", ""),
                job.get("url", ""),
            ]
            aligns = ["center", "left", "left", "left", "left",
                      "center", "center", "center", "left", "left"]
            for col, (val, aln) in enumerate(zip(values, aligns), 1):
                _cell_style(ws.cell(row_idx, col), val, fill_color=fill, align=aln)

            if job.get("url"):
                cell = ws.cell(row_idx, 10)
                cell.hyperlink = job["url"]
                cell.font = Font(color=COLOR_ACCENT, underline="single", size=10)

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    def _build_summary(self, ws, jobs, profile):
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20

        def section_header(row, text):
            cell = ws.cell(row, 1, text)
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(f"A{row}:B{row}")
            ws.row_dimensions[row].height = 22

        def data_row(row, label, value, bold_val=False):
            lc = ws.cell(row, 1, label)
            lc.font = Font(bold=True, size=10)
            lc.fill = PatternFill("solid", fgColor=COLOR_SECTION)
            lc.alignment = Alignment(horizontal="left", vertical="center")
            lc.border = _thin_border()
            vc = ws.cell(row, 2, value)
            vc.font = Font(bold=bold_val, size=10)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.border = _thin_border()

        r = 1
        ws.cell(r, 1, "JobSearch Pro — Run Summary").font = Font(bold=True, size=14, color=COLOR_ACCENT)
        ws.merge_cells(f"A{r}:B{r}")
        r += 1
        ws.cell(r, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, size=9)
        ws.merge_cells(f"A{r}:B{r}")
        r += 2

        section_header(r, "Search Overview")
        r += 1
        data_row(r, "Profile", profile.get("name", ""), bold_val=True); r += 1
        data_row(r, "Total Jobs Found", len(jobs), bold_val=True); r += 1
        data_row(r, "Remote Jobs", sum(1 for j in jobs if "remote" in j.get("location","").lower())); r += 1
        data_row(r, "Hybrid Jobs", sum(1 for j in jobs if "hybrid" in j.get("location","").lower())); r += 1
        data_row(r, "Work Type Filter", profile.get("work_type", "remote").capitalize()); r += 1
        data_row(r, "Salary Min", f"${profile.get('salary_min', 0):,}" if profile.get("salary_min") else "Not set"); r += 1

        r += 1
        section_header(r, "Results by Source")
        r += 1
        sources = {}
        for j in jobs:
            s = j.get("source", "unknown")
            sources[s] = sources.get(s, 0) + 1
        for src, count in sorted(sources.items(), key=lambda x: -x[1]):
            data_row(r, src.capitalize(), count); r += 1

        r += 1
        section_header(r, "Top Companies")
        r += 1
        companies = {}
        for j in jobs:
            c = j.get("company", "")
            if c:
                companies[c] = companies.get(c, 0) + 1
        for co, cnt in sorted(companies.items(), key=lambda x: -x[1])[:10]:
            data_row(r, co, cnt); r += 1

    def _build_config(self, ws, profile):
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 60

        _header_style(ws.cell(1, 1), "Setting")
        _header_style(ws.cell(1, 2), "Value")
        ws.row_dimensions[1].height = 28

        rows = [
            ("Profile Name", profile.get("name", "")),
            ("Work Type", profile.get("work_type", "remote").capitalize()),
            ("Salary Minimum", f"${profile.get('salary_min', 0):,}"),
            ("Sources", ", ".join(profile.get("sources", []))),
            ("Number of Queries", len(profile.get("queries", []))),
        ]
        for r, (k, v) in enumerate(rows, 2):
            fill = COLOR_SECTION if r % 2 == 0 else "FFFFFF"
            _cell_style(ws.cell(r, 1), k, fill_color=fill, bold=True)
            _cell_style(ws.cell(r, 2), str(v), fill_color=fill)

        r = len(rows) + 3
        ws.cell(r, 1, "Search Queries").font = Font(bold=True, size=11, color=COLOR_ACCENT)
        ws.merge_cells(f"A{r}:B{r}")
        r += 1

        _header_style(ws.cell(r, 1), "#")
        _header_style(ws.cell(r, 2), "Query")
        r += 1

        for i, q in enumerate(profile.get("queries", []), 1):
            text = q["query"] if isinstance(q, dict) else q
            fill = COLOR_ALT_ROW if i % 2 == 0 else "FFFFFF"
            _cell_style(ws.cell(r, 1), i, fill_color=fill, align="center")
            _cell_style(ws.cell(r, 2), text, fill_color=fill)
            r += 1
